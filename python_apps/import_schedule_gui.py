import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
import tkinter.messagebox as messagebox
import tkinter.filedialog as filedialog
import openpyxl
import mysql.connector
import datetime
import sys

# 設定
ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("スケジュール取り込みツール")
        self.geometry("500x350")
        
        self.file_path = None
        self.sheet_names = []
        
        # UI構成
        self.label = ctk.CTkLabel(self, text="エクセルファイルを選択してください", font=("Arial", 16))
        self.label.pack(pady=20)
        
        self.btn_select = ctk.CTkButton(self, text="ファイルを選択", command=self.select_file)
        self.btn_select.pack(pady=10)
        
        self.lbl_file = ctk.CTkLabel(self, text="未選択", text_color="gray")
        self.lbl_file.pack(pady=5)
        
        self.combo_sheet = ctk.CTkComboBox(self, values=["---"], state="disabled")
        self.combo_sheet.pack(pady=10)
        
        self.btn_run = ctk.CTkButton(self, text="インポート実行", command=self.run_import, state="disabled")
        self.btn_run.pack(pady=20)

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="エクセルファイルを開く",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.file_path = file_path
            self.lbl_file.configure(text=file_path)
            self.load_sheets()

    def load_sheets(self):
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=True)
            self.sheet_names = wb.sheetnames
            wb.close()
            
            self.combo_sheet.configure(values=self.sheet_names, state="normal")
            # "26.8" などのスケジュールっぽいシートがあればデフォルト選択
            for s in self.sheet_names:
                if '26' in s or '27' in s or 'スケジュール' in s:
                    self.combo_sheet.set(s)
                    break
            else:
                self.combo_sheet.set(self.sheet_names[0])
                
            self.btn_run.configure(state="normal")
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルの読み込みに失敗しました:\n{e}")

    def run_import(self):
        if not self.file_path:
            return
            
        sheet_name = self.combo_sheet.get()
        if not sheet_name or sheet_name == "---":
            messagebox.showwarning("警告", "シートを選択してください。")
            return
            
        try:
            # データベース接続
            conn = mysql.connector.connect(
                host='192.168.10.101',
                user='work_user',
                password='work1234',
                database='work_manager_db'
            )
            cursor = conn.cursor(dictionary=True)
            
            # マッピングの取得
            cursor.execute("SELECT original_name, mapped_name FROM m_model_mapping")
            model_map = {row['original_name']: row['mapped_name'] for row in cursor.fetchall()}
            
            cursor.execute("SELECT original_name, mapped_name FROM m_maker_mapping")
            maker_map = {row['original_name']: row['mapped_name'] for row in cursor.fetchall()}
            
            # Excel読み込み
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb[sheet_name]
            
            # 日付行（2行目）の取得
            dates = {}
            for col_idx in range(9, ws.max_column + 1):  # I列(9)以降と想定
                cell_val = ws.cell(row=2, column=col_idx).value
                if isinstance(cell_val, datetime.datetime):
                    dates[col_idx] = cell_val.date()
                    
            if not dates:
                messagebox.showerror("エラー", "2行目に日付データが見つかりませんでした。")
                return
                
            min_date = min(dates.values())
            max_date = max(dates.values())
            
            # 既存データの確認
            cursor.execute('''
                SELECT COUNT(*) as cnt 
                FROM t_schedules 
                WHERE target_date BETWEEN %s AND %s
            ''', (min_date, max_date))
            result = cursor.fetchone()
            
            if result['cnt'] > 0:
                # ユーザーへの確認ポップアップ
                ans = messagebox.askyesno(
                    "上書き確認",
                    f"{min_date.strftime('%Y/%m/%d')} ～ {max_date.strftime('%Y/%m/%d')} の期間に、すでに予定が登録されています。\nエクセルの予定数で上書きして実行しますか？"
                )
                if not ans:
                    return
                    
            # 取り込み処理
            now = datetime.datetime.now()
            insert_count = 0
            update_count = 0
            
            last_orig_model = None
            skipped_models = set()
            
            for row_idx in range(3, ws.max_row + 1):
                orig_model = ws.cell(row=row_idx, column=2).value
                orig_maker = ws.cell(row=row_idx, column=3).value
                
                # セル結合対策：値があれば記憶し、なければ直前の値を引き継ぐ
                if orig_model is not None and str(orig_model).strip() != "":
                    last_orig_model = str(orig_model).strip()
                    
                orig_model_str = last_orig_model
                
                if not orig_model_str:
                    continue
                
                if orig_model_str in skipped_models:
                    continue
                    
                # 「小計」「合計」や集計表の項目が含まれる行はスキップ
                skip_words = ["小計", "合計", "ツール数", "スタンド数", "稼働時間", "試験", "映像", "改造", "初期化", "ツール", "検品", "出荷登録", "スタンド", "アダプタ", "ACコード", "結束", "通電"]
                if any(sw in orig_model_str for sw in skip_words):
                    continue
                    
                orig_maker_str = str(orig_maker).strip() if orig_maker else ""
                # 全角スラッシュを半角スラッシュに統一
                orig_maker_str = orig_maker_str.replace("／", "/")
                
                model_name = model_map.get(orig_model_str, orig_model_str)
                maker_name = maker_map.get(orig_maker_str, "")
                
                for col_idx, target_date in dates.items():
                    plan_val = ws.cell(row=row_idx, column=col_idx).value
                    if plan_val is None or str(plan_val).strip() == "":
                        continue
                        
                    try:
                        plan_count = int(float(str(plan_val).strip()))
                    except ValueError:
                        continue
                        
                    # マスター(m_models)の登録状況をチェック
                    cursor.execute('''
                        SELECT DISTINCT maker_abbr FROM m_models 
                        WHERE model_name = %s
                    ''', (model_name,))
                    master_makers = cursor.fetchall()
                    
                    # 完全に未知の機種か判定
                    if not master_makers:
                        # マスターに存在しない場合、ユーザーにダイアログで選択させる
                        selected_action = None
                        mapped_model_name = None
                        
                        def on_map():
                            nonlocal selected_action, mapped_model_name
                            selected_action = "map"
                            mapped_model_name = combo_models.get()
                            popup.destroy()
                            
                        def on_new():
                            nonlocal selected_action
                            selected_action = "new"
                            popup.destroy()
                            
                        def on_skip():
                            nonlocal selected_action
                            selected_action = "skip"
                            popup.destroy()
                            
                        popup = tk.Toplevel(self)
                        popup.title("未知の機種が見つかりました")
                        popup.geometry("500x350")
                        popup.transient(self)
                        popup.grab_set()
                        
                        tk.Label(popup, text=f"エクセル内に未知の機種名が見つかりました:\n\n「{orig_model_str}」\n\nどう処理しますか？", justify="left").pack(pady=10, padx=10)
                        
                        # 既存機種リストの取得 (MySQL 3065 エラー対策として DISTINCT + ORDER BY を GROUP BY + MIN() に変更)
                        cursor.execute("SELECT model_name FROM m_models GROUP BY model_name ORDER BY MIN(sort_order) ASC")
                        all_models = [r['model_name'] for r in cursor.fetchall()]
                        
                        tk.Label(popup, text="▼ 既存の機種に変換（マッピング）する場合 ▼").pack(pady=5)
                        combo_models = ttk.Combobox(popup, values=all_models, width=40, state="readonly")
                        if all_models:
                            combo_models.set(all_models[0])
                        combo_models.pack(pady=5)
                        tk.Button(popup, text="この機種に変換する", command=on_map, bg="lightblue").pack(pady=5)
                        
                        tk.Label(popup, text="▼ または ▼").pack(pady=10)
                        tk.Button(popup, text="新機種としてそのままマスターに追加する", command=on_new, bg="lightgreen").pack(pady=5)
                        
                        tk.Label(popup, text="▼ または ▼").pack(pady=5)
                        tk.Button(popup, text="この行をスキップする（取り込まない）", command=on_skip, bg="lightcoral").pack(pady=5)
                        
                        self.wait_window(popup)
                        
                        if selected_action == "map" and mapped_model_name:
                            # マッピングを保存
                            cursor.execute("INSERT INTO m_model_mapping (original_name, mapped_name) VALUES (%s, %s) ON DUPLICATE KEY UPDATE mapped_name = %s", 
                                           (orig_model_str, mapped_model_name, mapped_model_name))
                            conn.commit()
                            model_map[orig_model_str] = mapped_model_name
                            model_name = mapped_model_name
                            
                            # マスター情報を再取得
                            cursor.execute('SELECT DISTINCT maker_abbr FROM m_models WHERE model_name = %s', (model_name,))
                            master_makers = cursor.fetchall()
                            
                        elif selected_action == "new":
                            # 新規機種として m_model_mapping に自分自身をマッピングして記憶
                            cursor.execute("INSERT INTO m_model_mapping (original_name, mapped_name) VALUES (%s, %s) ON DUPLICATE KEY UPDATE mapped_name = %s", 
                                           (orig_model_str, orig_model_str, orig_model_str))
                            conn.commit()
                            model_map[orig_model_str] = orig_model_str
                            model_name = orig_model_str
                            # マスターには存在しないが、エラーを回避するためにダミーのメーカー情報をセット
                            master_makers = [{'maker_abbr': ''}]
                        elif selected_action == "skip":
                            skipped_models.add(orig_model_str)
                            break
                        else:
                            # 閉じるボタンなどでキャンセルされた場合はスキップ
                            skipped_models.add(orig_model_str)
                            break

                    # もしマスターに登録されているメーカーが空文字1種類だけなら、エクセルのメーカー指定を無視して空文字にする
                    if len(master_makers) == 1 and (master_makers[0]['maker_abbr'] is None or master_makers[0]['maker_abbr'] == ''):
                        maker_name = ""
                        
                    # すでにDBにあるかチェック
                    cursor.execute('''
                        SELECT id FROM t_schedules 
                        WHERE target_date = %s AND model_name = %s AND maker_name = %s
                    ''', (target_date, model_name, maker_name))
                    existing = cursor.fetchone()
                    
                    if existing:
                        cursor.execute('''
                            UPDATE t_schedules 
                            SET plan_count = %s, updated_at = %s 
                            WHERE id = %s
                        ''', (plan_count, now, existing['id']))
                        update_count += 1
                    else:
                        cursor.execute('''
                            INSERT INTO t_schedules 
                            (target_date, model_name, maker_name, plan_count, actual_count, created_at, updated_at) 
                            VALUES (%s, %s, %s, %s, 0, %s, %s)
                        ''', (target_date, model_name, maker_name, plan_count, now, now))
                        insert_count += 1
                        
            # Flutterアプリ側に更新を通知するため、trackerを更新する
            cursor.execute("UPDATE data_update_tracker SET last_updated = CURRENT_TIMESTAMP WHERE id = 1")
            
            conn.commit()
            messagebox.showinfo("完了", f"データベースへのインポートが完了しました。\n新規追加: {insert_count}件\n上書き更新: {update_count}件")
            
        except Exception as e:
            messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}")
            if 'conn' in locals() and conn.is_connected():
                conn.rollback()
        finally:
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals() and conn.is_connected():
                conn.close()

if __name__ == "__main__":
    app = App()
    app.mainloop()
